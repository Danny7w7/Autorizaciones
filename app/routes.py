from flask import Blueprint, render_template, request, redirect, url_for, jsonify
from . import db
from .models import Authorization

bp = Blueprint("main", __name__)

@bp.route("/")
def index():
    authorizations = Authorization.query.all()
    return render_template("index.html", authorizations=authorizations)

@bp.route("/addAuthorization/", methods=["POST"])
def add():
    try:
        data = request.get_json()
        # Validar que exista el campo en el formulario
        authorization = data.get("autorizacion")
        if not authorization:
            return jsonify({"error": "El campo 'autorizacion' es obligatorio"}), 400

        # Verificar si ya existe en la base de datos
        existAuthorization = Authorization.query.filter_by(code=authorization).first()
        if existAuthorization:
            print('Esta autorizacion ya existe')
            return jsonify({"message": "Esta autorización ya existe"}), 400

        # Crear nuevo registro
        new = Authorization(code=authorization)
        db.session.add(new)
        db.session.commit()

        return jsonify({"message": "Autorización agregada correctamente"})

    except Exception as e:
        # Si ocurre cualquier error inesperado, revertimos la transacción
        db.session.rollback()
        return jsonify({"message": f"Ocurrió un error al procesar la solicitud: {e}"}), 500

    finally:
        # Cerramos la sesión de la base de datos
        db.session.close()



from openpyxl import load_workbook

@bp.route("/migrate/")
def migrate():
    # Ruta del archivo Excel
    ruta_archivo = "datos.xlsx"

    # Cargar el archivo
    wb = load_workbook(ruta_archivo)

    # Seleccionar la primera hoja
    hoja = wb.active

    # Iterar sobre todas las filas
    for fila in hoja.iter_rows(values_only=True):
        created_at = fila[0]
        code = fila[1]
        if created_at:
            dateNew = created_at
            print(dateNew)
        
        new = Authorization(code=code, created_at=dateNew)
        db.session.add(new)
        db.session.commit()